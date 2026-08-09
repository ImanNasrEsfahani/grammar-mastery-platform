from __future__ import annotations

from contextlib import contextmanager
from copy import deepcopy
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import hmac
import io
import json
from pathlib import PurePath
import secrets
import threading
from typing import Any, Iterator
import uuid

from .dedupe import classify_duplicates
from .normalization import normalize_row
from .validator import EXPECTED_COLUMNS, LookupCatalog, RowError, SCHEMA_VERSION, validate_row


PIPELINE_VERSION = "stage23-import-pipeline-v1.0.0"
MAX_FILE_BYTES = 20 * 1024 * 1024
MAX_ROWS = 1000


class PipelineError(Exception):
    def __init__(self, code: str, message: str, fields: dict[str, list[str]] | None = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.fields = fields or {}


def utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


@dataclass(frozen=True)
class ParsedFile:
    headers: list[str]
    rows: list[dict[str, Any]]


class InMemoryImportRepository:
    """Transactional reference adapter; PostgreSQL Patch 007 is the production shape."""

    def __init__(self) -> None:
        self.batches: dict[str, dict[str, Any]] = {}
        self.raw_files: dict[str, bytes] = {}
        self.rows: dict[str, list[dict[str, Any]]] = {}
        self.questions: dict[str, dict[str, Any]] = {}
        self.events: list[dict[str, Any]] = []
        self._lock = threading.RLock()

    @contextmanager
    def transaction(self) -> Iterator[None]:
        with self._lock:
            snapshot = deepcopy((self.batches, self.raw_files, self.rows, self.questions, self.events))
            try:
                yield
            except Exception:
                self.batches, self.raw_files, self.rows, self.questions, self.events = snapshot
                raise

    def existing_fingerprints(self) -> set[str]:
        return {str(row["fingerprint_sha256"]) for row in self.questions.values()}

    def existing_semantic_signatures(self) -> set[str]:
        return {str(row["semantic_signature_sha256"]) for row in self.questions.values()}


class ImportPipeline:
    def __init__(self, repository: InMemoryImportRepository | None = None, *, now=utcnow_iso) -> None:
        self.repository = repository or InMemoryImportRepository()
        self.now = now

    def _batch(self, batch_id: str) -> dict[str, Any]:
        try:
            return self.repository.batches[str(batch_id)]
        except KeyError as exc:
            raise PipelineError("IMPORT_BATCH_NOT_FOUND", "Import batch was not found.") from exc

    def _require_state(self, batch: dict[str, Any], *states: str) -> None:
        if batch["status"] not in states:
            raise PipelineError(
                "IMPORT_STATE_CONFLICT",
                f"Operation requires {', '.join(states)}; current state is {batch['status']}.",
            )

    def _event(self, batch_id: str, event_type: str, actor_id: str, details: dict[str, Any] | None = None) -> None:
        self.repository.events.append(
            {
                "event_id": str(uuid.uuid4()),
                "batch_id": str(batch_id),
                "event_type": event_type,
                "actor_id": str(actor_id),
                "details": deepcopy(details or {}),
                "occurred_at": self.now(),
            }
        )

    def _mark_failed(self, batch_id: str, actor_id: str, error: PipelineError) -> None:
        with self.repository.transaction():
            batch = self._batch(batch_id)
            batch["status"] = "FAILED"
            batch["failure_code"] = error.code
            batch["failed_at"] = self.now()
            self._event(batch_id, "FAILED", actor_id, {"code": error.code, "fields": deepcopy(error.fields)})

    def upload(
        self,
        *,
        file_name: str,
        content: bytes,
        actor_id: str,
        batch_id: str | None = None,
        schema_version: str = SCHEMA_VERSION,
    ) -> dict[str, Any]:
        batch_id = str(batch_id or uuid.uuid4())
        try:
            uuid.UUID(batch_id)
        except ValueError as exc:
            raise PipelineError("IMPORT_BATCH_ID_INVALID", "batch_id must be a UUID.") from exc
        if batch_id in self.repository.batches:
            raise PipelineError("IMPORT_BATCH_EXISTS", "batch_id already exists and cannot be overwritten.")
        suffix = PurePath(file_name).suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise PipelineError("IMPORT_FILE_TYPE_UNSUPPORTED", "Only CSV and XLSX files are accepted.")
        if not content or len(content) > MAX_FILE_BYTES:
            raise PipelineError("IMPORT_FILE_SIZE_INVALID", f"File must contain 1-{MAX_FILE_BYTES} bytes.")
        if schema_version != SCHEMA_VERSION:
            raise PipelineError("SCHEMA_VERSION_UNSUPPORTED", f"Use {SCHEMA_VERSION}.")
        raw_hash = hashlib.sha256(content).hexdigest()
        batch = {
            "batch_id": batch_id,
            "pipeline_version": PIPELINE_VERSION,
            "schema_version": schema_version,
            "file_name": PurePath(file_name).name,
            "file_type": suffix[1:].upper(),
            "raw_sha256": raw_hash,
            "raw_size_bytes": len(content),
            "status": "UPLOADED",
            "created_by": str(actor_id),
            "created_at": self.now(),
            "row_count": 0,
            "valid_count": 0,
            "invalid_count": 0,
            "duplicate_count": 0,
            "semantic_review_count": 0,
            "committed_count": 0,
            "preview_sha256": None,
            "confirmation_token_sha256": None,
        }
        with self.repository.transaction():
            self.repository.batches[batch_id] = batch
            self.repository.raw_files[batch_id] = bytes(content)
            self._event(batch_id, "UPLOADED", actor_id, {"raw_sha256": raw_hash, "size": len(content)})
        return deepcopy(batch)

    @staticmethod
    def _parse_csv(content: bytes) -> ParsedFile:
        try:
            text = content.decode("utf-8-sig", errors="strict")
        except UnicodeDecodeError as exc:
            raise PipelineError("IMPORT_ENCODING_INVALID", "CSV must be UTF-8.") from exc
        reader = csv.DictReader(io.StringIO(text, newline=""))
        headers = [str(header or "").lstrip("\ufeff").strip() for header in (reader.fieldnames or [])]
        rows: list[dict[str, Any]] = []
        for row_number, source in enumerate(reader, start=2):
            if None in source:
                raise PipelineError("IMPORT_CSV_SHAPE_INVALID", f"Row {row_number} has more cells than headers.")
            if not any(str(value or "").strip() for value in source.values()):
                continue
            source["_row_number"] = row_number
            rows.append(source)
        return ParsedFile(headers, rows)

    @staticmethod
    def _parse_xlsx(content: bytes) -> ParsedFile:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise PipelineError("XLSX_ADAPTER_UNAVAILABLE", "Install openpyxl 3.1+ for XLSX imports.") from exc
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            iterator = worksheet.iter_rows(values_only=True)
            first = next(iterator, None)
            headers = [str(value or "").lstrip("\ufeff").strip() for value in (first or [])]
            rows = []
            for row_number, values in enumerate(iterator, start=2):
                if not any(str(value or "").strip() for value in values):
                    continue
                if len(values) > len(headers) and any(value is not None for value in values[len(headers):]):
                    raise PipelineError("IMPORT_XLSX_SHAPE_INVALID", f"Row {row_number} has more cells than headers.")
                row = {name: values[index] if index < len(values) else None for index, name in enumerate(headers)}
                row["_row_number"] = row_number
                rows.append(row)
            workbook.close()
            return ParsedFile(headers, rows)
        except PipelineError:
            raise
        except Exception as exc:
            raise PipelineError("IMPORT_XLSX_INVALID", "XLSX workbook could not be parsed safely.") from exc

    def parse(self, batch_id: str, *, actor_id: str) -> dict[str, Any]:
        batch = self._batch(batch_id)
        self._require_state(batch, "UPLOADED")
        content = self.repository.raw_files[batch_id]
        try:
            parsed = self._parse_csv(content) if batch["file_type"] == "CSV" else self._parse_xlsx(content)
            if parsed.headers != list(EXPECTED_COLUMNS):
                missing = sorted(set(EXPECTED_COLUMNS) - set(parsed.headers))
                unknown = sorted(set(parsed.headers) - set(EXPECTED_COLUMNS))
                raise PipelineError(
                    "IMPORT_HEADER_INVALID",
                    "The header must match the frozen 46-column schema in exact order.",
                    {"missing": missing, "unknown": unknown},
                )
            if not parsed.rows or len(parsed.rows) > MAX_ROWS:
                raise PipelineError("IMPORT_ROW_COUNT_INVALID", f"A batch must contain 1-{MAX_ROWS} data rows.")
        except PipelineError as error:
            self._mark_failed(batch_id, actor_id, error)
            raise
        with self.repository.transaction():
            self.repository.rows[batch_id] = [normalize_row(row) | {"_row_number": row["_row_number"]} for row in parsed.rows]
            batch["row_count"] = len(parsed.rows)
            batch["status"] = "PARSED"
            batch["parsed_at"] = self.now()
            self._event(batch_id, "PARSED", actor_id, {"row_count": len(parsed.rows)})
        return deepcopy(batch)

    @staticmethod
    def _row_issue(row: dict[str, Any], issue: RowError | dict[str, Any]) -> None:
        row.setdefault("issues", []).append(issue.to_dict() if isinstance(issue, RowError) else deepcopy(issue))

    def _preview_payload(self, batch: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any]:
        errors = [deepcopy(issue) for row in rows for issue in row.get("issues", [])]
        return {
            "batch_id": batch["batch_id"],
            "schema_version": batch["schema_version"],
            "pipeline_version": batch["pipeline_version"],
            "raw_sha256": batch["raw_sha256"],
            "row_count": len(rows),
            "valid_count": sum(row["disposition"] == "VALID" for row in rows),
            "invalid_count": sum(row["disposition"] == "INVALID" for row in rows),
            "duplicate_count": sum(row["disposition"] == "DUPLICATE" for row in rows),
            "semantic_review_count": sum(row["disposition"] == "SEMANTIC_REVIEW" for row in rows),
            "committable": all(row["disposition"] == "VALID" for row in rows),
            "errors": errors,
            "rows": [
                {
                    "row_number": row["_row_number"],
                    "external_id": row.get("external_id"),
                    "disposition": row["disposition"],
                    "fingerprint_sha256": row["fingerprint_sha256"],
                    "duplicate_classification": row["duplicate_classification"],
                }
                for row in rows
            ],
        }

    def _seal_preview(self, batch: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any]:
        payload = self._preview_payload(batch, rows)
        canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        preview_hash = hashlib.sha256(canonical).hexdigest()
        token = secrets.token_urlsafe(32)
        batch.update(
            {
                "valid_count": payload["valid_count"],
                "invalid_count": payload["invalid_count"],
                "duplicate_count": payload["duplicate_count"],
                "semantic_review_count": payload["semantic_review_count"],
                "preview_sha256": preview_hash,
                "confirmation_token_sha256": hashlib.sha256(token.encode("utf-8")).hexdigest(),
                "status": "PREVIEW_READY",
                "previewed_at": self.now(),
            }
        )
        payload["preview_sha256"] = preview_hash
        payload["confirmation_token"] = token
        return payload

    def validate_and_preview(self, batch_id: str, *, actor_id: str, lookups: LookupCatalog) -> dict[str, Any]:
        batch = self._batch(batch_id)
        self._require_state(batch, "PARSED")
        rows = self.repository.rows[batch_id]
        duplicate_results = classify_duplicates(
            rows,
            existing_fingerprints=self.repository.existing_fingerprints(),
            existing_semantic_signatures=self.repository.existing_semantic_signatures(),
        )
        with self.repository.transaction():
            for row, duplicate in zip(rows, duplicate_results, strict=True):
                validation_issues = validate_row(row, int(row["_row_number"]), lookups)
                row["issues"] = []
                for issue in validation_issues:
                    self._row_issue(row, issue)
                row.update(duplicate)
                row["duplicate_classification"] = duplicate.pop("classification")
                classification = row["duplicate_classification"]
                if classification.startswith("EXACT"):
                    self._row_issue(
                        row,
                        {
                            "row_number": row["_row_number"],
                            "field": "fingerprint_sha256",
                            "code": "DUPLICATE_EXACT",
                            "message": "Exact/fingerprint duplicate cannot be committed.",
                            "severity": "ERROR",
                        },
                    )
                    row["disposition"] = "DUPLICATE"
                elif "SEMANTIC" in classification:
                    self._row_issue(
                        row,
                        {
                            "row_number": row["_row_number"],
                            "field": "stem",
                            "code": "SEMANTIC_REVIEW_REQUIRED",
                            "message": "A near duplicate requires an explicit reviewed decision.",
                            "severity": "WARNING",
                        },
                    )
                    row["disposition"] = "SEMANTIC_REVIEW"
                elif row["issues"]:
                    row["disposition"] = "INVALID"
                else:
                    row["disposition"] = "VALID"
            report = self._seal_preview(batch, rows)
            self._event(
                batch_id,
                "PREVIEWED",
                actor_id,
                {key: report[key] for key in ("row_count", "valid_count", "invalid_count", "duplicate_count", "semantic_review_count", "preview_sha256")},
            )
        return deepcopy(report)

    def resolve_semantic(
        self,
        batch_id: str,
        *,
        row_number: int,
        decision: str,
        reason: str,
        actor_id: str,
    ) -> dict[str, Any]:
        batch = self._batch(batch_id)
        self._require_state(batch, "PREVIEW_READY")
        if decision not in {"ACCEPT_DISTINCT", "REJECT_DUPLICATE"} or not reason.strip():
            raise PipelineError("SEMANTIC_DECISION_INVALID", "Decision and non-empty reason are required.")
        row = next((item for item in self.repository.rows[batch_id] if item["_row_number"] == row_number), None)
        if row is None or row.get("disposition") != "SEMANTIC_REVIEW":
            raise PipelineError("SEMANTIC_REVIEW_NOT_FOUND", "Row is not awaiting semantic review.")
        with self.repository.transaction():
            row["semantic_resolution"] = {"decision": decision, "reason": reason.strip(), "actor_id": str(actor_id), "resolved_at": self.now()}
            row["issues"] = [issue for issue in row["issues"] if issue["code"] != "SEMANTIC_REVIEW_REQUIRED"]
            if decision == "ACCEPT_DISTINCT":
                row["disposition"] = "INVALID" if any(issue["severity"] == "ERROR" for issue in row["issues"]) else "VALID"
            else:
                self._row_issue(row, {"row_number": row_number, "field": "stem", "code": "DUPLICATE_SEMANTIC_REJECTED", "message": reason.strip(), "severity": "ERROR"})
                row["disposition"] = "DUPLICATE"
            report = self._seal_preview(batch, self.repository.rows[batch_id])
            self._event(batch_id, "SEMANTIC_REVIEW_RESOLVED", actor_id, {"row_number": row_number, "decision": decision, "reason": reason.strip()})
        return deepcopy(report)

    def commit(self, batch_id: str, *, confirmation_token: str, actor_id: str) -> dict[str, Any]:
        batch = self._batch(batch_id)
        self._require_state(batch, "PREVIEW_READY")
        supplied = hashlib.sha256(confirmation_token.encode("utf-8")).hexdigest()
        if not hmac.compare_digest(supplied, str(batch["confirmation_token_sha256"])):
            raise PipelineError("IMPORT_CONFIRMATION_INVALID", "Confirmation token is invalid or stale.")
        rows = self.repository.rows[batch_id]
        if not rows or any(row["disposition"] != "VALID" for row in rows):
            raise PipelineError("IMPORT_NOT_COMMITTABLE", "Resolve or remove every invalid, duplicate and review row, then preview again.")
        with self.repository.transaction():
            if any(row["fingerprint_sha256"] in self.repository.existing_fingerprints() for row in rows):
                raise PipelineError("IMPORT_PREVIEW_STALE", "Question inventory changed after preview; run a new batch preview.")
            seen_keys = {(str(question["external_id"]), int(question["question_revision"])) for question in self.repository.questions.values()}
            for row in rows:
                key = (str(row["external_id"]), int(row["question_revision"]))
                if key in seen_keys:
                    raise PipelineError("IMPORT_EXTERNAL_REVISION_CONFLICT", "external_id/revision already exists.")
                question_id = str(uuid.uuid4())
                self.repository.questions[question_id] = {
                    "id": question_id,
                    "question_uid": str(uuid.uuid4()),
                    "external_id": row["external_id"],
                    "question_revision": int(row["question_revision"]),
                    "status": "DRAFT",
                    "fingerprint_sha256": row["fingerprint_sha256"],
                    "semantic_signature_sha256": row["semantic_signature_sha256"],
                    "import_batch_id": batch_id,
                    "source_row_number": row["_row_number"],
                    "payload": {key: deepcopy(row[key]) for key in EXPECTED_COLUMNS},
                    "created_at": self.now(),
                }
                row["committed_question_id"] = question_id
                seen_keys.add(key)
            batch["status"] = "COMMITTED"
            batch["committed_count"] = len(rows)
            batch["committed_at"] = self.now()
            batch["confirmation_token_sha256"] = None
            self._event(batch_id, "COMMITTED", actor_id, {"committed_count": len(rows), "preview_sha256": batch["preview_sha256"]})
            self._post_check(batch_id)
        return {"batch_id": batch_id, "status": "COMMITTED", "committed_count": len(rows), "preview_sha256": batch["preview_sha256"]}

    def _post_check(self, batch_id: str) -> None:
        batch = self._batch(batch_id)
        linked = [question for question in self.repository.questions.values() if question["import_batch_id"] == batch_id]
        if len(linked) != batch["committed_count"] or any(question["status"] != "DRAFT" for question in linked):
            raise PipelineError("IMPORT_POST_CHECK_FAILED", "Committed count or DRAFT-only invariant failed; transaction was rolled back.")

    def rollback(self, batch_id: str, *, actor_id: str, reason: str) -> dict[str, Any]:
        batch = self._batch(batch_id)
        self._require_state(batch, "COMMITTED")
        if not reason.strip():
            raise PipelineError("ROLLBACK_REASON_REQUIRED", "A rollback reason is required.")
        linked = [question for question in self.repository.questions.values() if question["import_batch_id"] == batch_id]
        blocked = [question["id"] for question in linked if question["status"] != "DRAFT" or question.get("has_downstream_references")]
        if blocked:
            raise PipelineError(
                "ROLLBACK_REQUIRES_RETIREMENT_WORKFLOW",
                "Rollback is limited to untouched DRAFT rows; changed/referenced questions require the audited retirement workflow.",
                {"question_ids": blocked},
            )
        with self.repository.transaction():
            for question in linked:
                del self.repository.questions[question["id"]]
            batch["status"] = "ROLLED_BACK"
            batch["rolled_back_count"] = len(linked)
            batch["rolled_back_at"] = self.now()
            batch["rollback_reason"] = reason.strip()
            self._event(batch_id, "ROLLED_BACK", actor_id, {"rolled_back_count": len(linked), "reason": reason.strip()})
        return {"batch_id": batch_id, "status": "ROLLED_BACK", "rolled_back_count": len(linked)}

    def batch_audit(self, batch_id: str) -> dict[str, Any]:
        batch = deepcopy(self._batch(batch_id))
        batch.pop("confirmation_token_sha256", None)
        return {
            "batch": batch,
            "events": [deepcopy(event) for event in self.repository.events if event["batch_id"] == batch_id],
            "raw_retained": batch_id in self.repository.raw_files,
            "raw_sha256_verified": hashlib.sha256(self.repository.raw_files[batch_id]).hexdigest() == batch["raw_sha256"],
        }
